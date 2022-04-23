# coding:utf-8
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import config as cf
from bs4 import BeautifulSoup  # BeautifulSoup をインポート
from openpyxl import Workbook
import subprocess
from openpyxl.styles import Alignment    # Alignmentクラスをインポート
from openpyxl.styles import Border, Side, PatternFill, Font
import xlwings as xw

# ブラウザを開く。
driver = webdriver.Chrome(executable_path=cf.CHROMEDRIVER)
driver.get("https://www.ricoh.co.jp/mfp/color/im-C6000-C5500-C4500-C3500-C3000-C2500/spec.html")
content = driver.page_source
BS = BeautifulSoup(content, "html.parser")
# テーブル"wp-block-table is-style-stripes"を指定
table = BS.findAll("table", {"class": "nml nml_scroll"})[0]
rows = table.findAll("tr")  # テーブル中<tr>要素の内容を抽出
print(rows)  # 抽出したHTMLドキュメントを検証
time.sleep(1)

wb = Workbook()
mk_name = 'test.xlsx'
wb.save(mk_name)
# ワークブックの読み込み
wb = load_workbook(mk_name)
ws = wb.active
row_count = 0
side1 = Side(style='thin', color='000000')
for row in rows:
    csvRow = []
    row_count = row_count + 1
    for cell in row.findAll(['td', 'th']):  # tdとth要素をループでファイルに書き込む
        csvRow.append(cell.get_text().replace(" ", ""))
    if csvRow[0] == "\xa0" :
        csvRow[0] == ""
    for idx in range(1,len(csvRow)+1):
        ws.cell(row=row_count, column=idx,value=csvRow[idx-1]).alignment = Alignment(wrapText=True) #折り返し設定
        ws.cell(row=row_count, column=idx).border = Border(left=side1, right=side1, top=side1, bottom=side1)
        if (not(row_count == 1 and idx == 1)) and (row_count == 1 or idx == 1):
            fill = PatternFill(patternType="solid", fgColor="D1FE7B")
            ws.cell(row=row_count, column=idx).fill = fill


ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 38
ws.column_dimensions['D'].width = 38
ws.column_dimensions['E'].width = 38
# ws['A2:A29'].styles.PatternFill(fgColor='FF0000', bgColor='FF0000')
# ws['A2:A29'].styles.PatternFill(patternType='solid',fgColor='FF0000', bgColor='FF0000')
# ws['A1':'E29'].alignment = Alignment(wrapText=True)
wb.save(mk_name)
EXCEL = r'C:\Users\chopp\Project\Python_Project\test.xlsx'
subprocess.Popen(['start', EXCEL], shell=True)
time.sleep(10)
