from openpyxl import Workbook
import subprocess
import time

# ワークブックの新規作成と保存
wb = Workbook()
mk_name = 'test.xlsx'
wb.save(mk_name)

# ワークブックの読み込み
from openpyxl import load_workbook

wb = load_workbook(mk_name)

# ワークシートの選択
ws = wb['Sheet']  # ワークシートを指定
ws = wb.active  # アクティブなワークシートを選択

# ワークシートの作成
wb.create_sheet('Login')
wb.create_sheet('Menu1')
wb.create_sheet('Menu2')
wb.create_sheet('Menu3')

wss =[]
# ワークシートの列挙
for sheet in wb:
    wss.append(sheet)

# 1シート名のセルに書き込み
wss[0]['A10'] = 'Hello from Python'
wb.save(mk_name)  # overwrite test.xlsx
# 2シート目のセルに書き込み
wss[1]['A5'] = 'Printing'
wb.save(mk_name)  # overwrite test.xlsx

# cellメソッドでセルに書き込み
ws.cell(row=1, column=1).value = 1
ws.cell(row=1, column=1, value=2)

for idx in range(1, 4):
    ws.cell(row=1, column=idx, value=idx)
    ws.cell(row=2, column=idx, value=-idx)

# ワークシートの全セルを反復（行ごと）
for row in ws.iter_rows():
    for cell in row:
        print(cell.value)
# 出力結果：
#1 2 3 -1 -2 -3

for row in ws.rows:
    for cell in row:
        print(cell.value)

min_row = ws.min_row
max_row = ws.max_row
min_col = ws.min_column
max_col = ws.max_column

for row in ws.iter_rows(min_row=min_row, max_row=max_row,min_col=min_col, max_col=max_col):
    for cell in row:
        print(cell.value)

# セルの値を取得する
for row in ws.iter_rows(values_only=True):
    for value in row:
        print(value)

# ワークシートの全セルを反復（列ごと）
for column in ws.iter_cols(values_only=True):
    print(column)

for column in ws.columns:
    for cell in column:
        print(cell.value)

# セル範囲の取得
rng = ws['A1':'C2']
ws.cell(row=6, column=1).value = "COPY[A1:C2]↓"

values = []
for row in rng:
    tmp = []
    for col in row:
        tmp.append(col.value)
    values.append(tmp)
#values = [[col.value for col in row] for row in rng]
print(values)

# セル範囲への代入
from openpyxl.cell.cell import Cell

def get_shape(rng):
    return (len(rng), len(rng[0]))

def assign2range(dst, src):
    dst_shape = get_shape(dst)
    src_shape = get_shape(src)
    if src_shape != dst_shape:
        raise ValueError('shapes of arguments not match')

    for d, s in zip(dst, src):  # dst／src: Cellまたは値のタプルを格納するタプル
        for t, v in zip(d, s):  # d／s: Cellまたは値を格納するタプル
            if isinstance(v, Cell):
                v = v.value
            t.value = v

values = [
    [1, 2, 3], [4, 5, 6]
]

assign2range(ws['A4':'C5'], values)
assign2range(ws['A9:C9'], [[100, 200, 300]])
assign2range(ws['A7':'C8'], ws['A1':'C2'])

#合計
wss[0]['A11'] = 'Sum↓'
sum = 0
for sum_row in range(4,6):
    sum += ws.cell(row=sum_row, column=1).value
wss[0]['A12'] = sum
    # wb.save('./Excelサンプル.xlsx')

wb.save(mk_name)

EXCEL = r'C:\Users\chopp\Project\Python_Project\test.xlsx'
subprocess.Popen(['start', EXCEL], shell=True)
time.sleep(10)
